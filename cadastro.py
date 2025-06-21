import customtkinter as ctk
from tkinter import *
from tkinter import messagebox
import openpyxl
import pathlib
import os
from openpyxl import Workbook


ctk.set_appearance_mode('System')
ctk.set_default_color_theme('blue')


class App(ctk.CTk):

    def abrir_planilha(self):
        caminho = os.path.abspath('Clientes.xlsx')
        os.startfile(caminho)

    def __init__(self):
        super().__init__()
        self.layout_config()
        self.aparencia()
        self.todo_sistema()

    def layout_config(self):
        self.title('Sistema de gestão de Cliente')
        self.geometry('700x500')

    def aparencia(self):
        lb_apc = ctk.CTkLabel(self, text='Tema', bg_color='transparent', text_color=['#000', "#fff"])
        lb_apc.place(x=50, y=430)

        apt_apc = ctk.CTkOptionMenu(self, values=['light', 'Dark', 'System'], command=self.change_apc)
        apt_apc.place(x=50, y=460)

    def todo_sistema(self):
       
        frame = ctk.CTkFrame(self, width=700, height=50, corner_radius=0, bg_color='teal', fg_color='teal')
        frame.place(x=0, y=12)

        title = ctk.CTkLabel(frame, text='Sistema de Gestão de Clientes', font=('Century Gothic bold', 24), text_color='#fff')
        title.place(relx=0.5, rely=0.5, anchor='center')

        span = ctk.CTkLabel(self, text='Por favor, preencha todos os campos do formulário',
                            font=('Century Gothic bold', 16), text_color=['#000', '#fff'])
        span.place(x=50, y=70)
        
        self.name_value = StringVar()
        self.contact_value = StringVar()
        self.age_value = StringVar()
        self.address_value = StringVar()

        
        ctk.CTkLabel(self, text='Nome:', font=('Century Gothic bold', 16), text_color=['#000', '#fff']).place(x=40, y=120)
        ctk.CTkLabel(self, text='Contato:', font=('Century Gothic bold', 16), text_color=['#000', '#fff']).place(x=40, y=170)
        ctk.CTkLabel(self, text='Idade:', font=('Century Gothic bold', 16), text_color=['#000', '#fff']).place(x=40, y=220)
        ctk.CTkLabel(self, text='Gênero:', font=('Century Gothic bold', 16), text_color=['#000', '#fff']).place(x=40, y=270)
        ctk.CTkLabel(self, text='Endereço:', font=('Century Gothic bold', 16), text_color=['#000', '#fff']).place(x=40, y=320)
        ctk.CTkLabel(self, text='Observações:', font=('Century Gothic bold', 16), text_color=['#000', '#fff']).place(x=40, y=370)

        
        self.name_entry = ctk.CTkEntry(self, width=350, textvariable=self.name_value, font=('Century Gothic bold', 16), fg_color='transparent')
        self.name_entry.place(x=150, y=120)

        self.contact_entry = ctk.CTkEntry(self, width=200, textvariable=self.contact_value, font=('Century Gothic bold', 16), fg_color='transparent')
        self.contact_entry.place(x=150, y=170)

        self.age_entry = ctk.CTkEntry(self, width=150, textvariable=self.age_value, font=('Century Gothic bold', 16), fg_color='transparent')
        self.age_entry.place(x=150, y=220)

        self.gender_combobox = ctk.CTkComboBox(self, values=['Masculino', 'Feminino'], font=('Century Gothic bold', 14))
        self.gender_combobox.set('Masculino')
        self.gender_combobox.place(x=150, y=270)

        self.address_entry = ctk.CTkEntry(self, width=300, textvariable=self.address_value, font=('Century Gothic bold', 16), fg_color='transparent')
        self.address_entry.place(x=150, y=320)

        self.obs_entry = ctk.CTkTextbox(self, width=500, height=80, font=('Arial', 16), border_color="#aaa", border_width=2, fg_color='transparent')
        self.obs_entry.place(x=150, y=370)

        
        ctk.CTkButton(self, text='ABRIR PLANILHA', command=self.abrir_planilha, fg_color='#0066cc', hover_color='#004c99').place(x=100, y=460)
        ctk.CTkButton(self, text='SALVAR DADOS', command=self.submit, fg_color='#151', hover_color='#131').place(x=300, y=460)
        ctk.CTkButton(self, text='LIMPAR CAMPOS', command=self.clear, fg_color='#555', hover_color='#333').place(x=500, y=460)

        
        ficheiro = pathlib.Path('Clientes.xlsx')
        if not ficheiro.exists():
            wb = Workbook()
            folha = wb.active
            folha['A1'] = 'Nome completo'
            folha['B1'] = 'Contato'
            folha['C1'] = 'Idade'
            folha['D1'] = 'Gênero'
            folha['E1'] = 'Endereço'
            folha['F1'] = 'Observações'
            wb.save('Clientes.xlsx')

    def submit(self):
        
        name = self.name_value.get()
        contact = self.contact_value.get()
        age = self.age_value.get()
        gender = self.gender_combobox.get()
        address = self.address_value.get()
        obs = self.obs_entry.get(0.0, END)

       
        ficheiro = openpyxl.load_workbook('Clientes.xlsx')
        folha = ficheiro.active
        row = folha.max_row + 1
        folha.cell(column=1, row=row, value=name)
        folha.cell(column=2, row=row, value=contact)
        folha.cell(column=3, row=row, value=age)
        folha.cell(column=4, row=row, value=gender)
        folha.cell(column=5, row=row, value=address)
        folha.cell(column=6, row=row, value=obs)
        ficheiro.save('Clientes.xlsx')

       
        messagebox.showinfo('Sistema', 'Dados salvos com sucesso!')

    def clear(self):
        self.name_value.set('')
        self.contact_value.set('')
        self.age_value.set('')
        self.address_value.set('')
        self.obs_entry.delete(0.0, END)

    def change_apc(self, nova_aparencia):
        ctk.set_appearance_mode(nova_aparencia)


if __name__ == '__main__':
    app = App()
    app.mainloop()
